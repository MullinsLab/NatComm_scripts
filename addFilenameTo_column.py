import os
from openpyxl import load_workbook

# Use the current directory (where the script is located)
directory = os.getcwd()

# Loop through all Excel files in the directory
for filename in os.listdir(directory):
    if filename.endswith(('.xlsx', '.xls')):  # Process only Excel files
        file_path = os.path.join(directory, filename)
        
        # Load the workbook
        wb = load_workbook(file_path)
        ws = wb.active  # Use the active sheet (or specify: wb['Sheet1'])
        
        # Add filename to cell A1 (or change to desired cell)
        ws['A1'] = filename
        
        # Save the modified workbook
        wb.save(file_path)
        print(f"Processed: {filename}")

print("Batch processing complete!")